# -*- coding: utf-8 -*-
"""
merge_worklogs_with_rates.py  (v5: strict preserve of 'Ссылка на задачу', no project column)

- НЕ добавляет "Ссылка на проект".
- Сохраняет столбец "Ссылка на задачу" ровно как в исходнике:
  * переносит текст ячейки;
  * переносит hyperlink.target, если он был (если не было — не добавляет).
- Совместимо с Python 3.7–3.9.
"""

import os, re
from typing import Optional, Union, Tuple
from pathlib import Path
from datetime import datetime

import pandas as pd
import numpy as np
from openpyxl import load_workbook

# ====== НАСТРОЙКИ ======
WL_DIR = r"C:\Users\zaytsev_ra2\PycharmProjects\TCO\WL"
RP_RESULT_DIR = r"C:\Users\zaytsev_ra2\PycharmProjects\TCO\Data\Resource plan\Результат"
SHEET_NAME_OUT = "worklogs_with_rates"
TASK_LINK_COL = "Ссылка на задачу"


def get_latest_file(folder: Union[str, os.PathLike], ext: Tuple[str, ...] = (".xlsx", ".xls")) -> Optional[Path]:
    folder = Path(folder)
    if not folder.exists():
        return None
    files = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in ext]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def to_number(val):
    if pd.isna(val):
        return np.nan
    if isinstance(val, (int, float, np.number)):
        return float(val)
    s = str(val).replace("\xa0", " ").strip().replace(" ", "").replace(",", ".")
    s = re.sub(r"[^\d\.\-]", "", s)
    try:
        return float(s)
    except ValueError:
        return np.nan


def pick_rate(series: pd.Series):
    clean = series.dropna()
    if clean.empty:
        return np.nan
    mode_vals = clean.mode()
    if not mode_vals.empty:
        return float(mode_vals.iloc[0])
    return float(clean.iloc[0])


def build_rates_table(rp_df: pd.DataFrame):
    if "Сотрудник" not in rp_df.columns:
        raise KeyError('В ресурсном плане нет колонки "Сотрудник"')
    if "Ставка" not in rp_df.columns:
        raise KeyError('В ресурсном плане нет колонки "Ставка"')
    tmp = rp_df[["Сотрудник", "Ставка"]].copy()
    tmp["Ставка"] = tmp["Ставка"].apply(to_number)
    conflicts = (
        tmp.dropna(subset=["Ставка"])
           .groupby("Сотрудник")["Ставка"]
           .nunique()
           .reset_index(name="unique_rates_count")
    )
    conflicts = conflicts[conflicts["unique_rates_count"] > 1]
    rates = (
        tmp.groupby("Сотрудник", as_index=False)["Ставка"]
           .agg(pick_rate)
           .rename(columns={"Ставка": "Ставка, ₽/ч"})
    )
    return rates, conflicts


def copy_original_task_links(src_wl: Path, dst_xlsx: Path, dst_sheet: Optional[str] = SHEET_NAME_OUT,
                             task_col: str = TASK_LINK_COL):
    """Копирует текст и гиперссылки из исходного WL в результат, 1:1, по порядку строк."""
    src_wb = load_workbook(src_wl, read_only=False, data_only=True)
    src_ws = src_wb.active
    src_header = [c.value for c in next(src_ws.iter_rows(min_row=1, max_row=1))]
    try:
        src_task_idx = src_header.index(task_col) + 1
    except ValueError:
        src_task_idx = None

    dst_wb = load_workbook(dst_xlsx)
    ws = dst_wb[dst_sheet] if dst_sheet else dst_wb.active
    dst_header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # гарантируем колонку в выходе
    if task_col in dst_header:
        dst_task_idx = dst_header.index(task_col) + 1
    else:
        ws.cell(row=1, column=len(dst_header) + 1, value=task_col)
        dst_task_idx = len(dst_header) + 1
        dst_header.append(task_col)

    max_rows = min(src_ws.max_row, ws.max_row)
    for r in range(2, max_rows + 1):
        src_text, src_target = None, None
        if src_task_idx:
            c_src = src_ws.cell(row=r, column=src_task_idx)
            src_text = c_src.value
            if c_src.hyperlink is not None:
                try:
                    src_target = c_src.hyperlink.target
                except Exception:
                    src_target = None
        c_dst = ws.cell(row=r, column=dst_task_idx)
        c_dst.value = src_text
        # если в исходнике не было гиперссылки — не добавляем
        if src_target:
            c_dst.hyperlink = src_target
            c_dst.style = "Hyperlink"
        else:
            c_dst.hyperlink = None

    dst_wb.save(dst_xlsx)
    src_wb.close()


def main(wl_path: Optional[Path] = None, rp_path: Optional[Path] = None) -> Path:
    if wl_path is None:
        wl_path = get_latest_file(WL_DIR)
    if rp_path is None:
        rp_path = get_latest_file(RP_RESULT_DIR)

    if wl_path is None:
        raise FileNotFoundError("Не найден Excel-файл в папке WL: {}".format(WL_DIR))
    if rp_path is None:
        raise FileNotFoundError("Не найден Excel-файл в папке ресурсного плана: {}".format(RP_RESULT_DIR))

    print("[i] WL:", wl_path)
    print("[i] RP:", rp_path)

    wl = pd.read_excel(wl_path)
    rp = pd.read_excel(rp_path)

    if "Сотрудник" not in wl.columns:
        raise KeyError('В worklogs нет колонки "Сотрудник"')
    rates, conflicts = build_rates_table(rp)

    out_dir = Path(wl_path).parent
    ts = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")

    # merge, сохраняя порядок WL
    merged = wl.merge(rates, on="Сотрудник", how="left")
    if "Часы факт" in merged.columns:
        merged["Стоимость факт, ₽"] = merged["Часы факт"].apply(to_number) * merged["Ставка, ₽/ч"]
    else:
        merged["Стоимость факт, ₽"] = np.nan

    # порядок колонок
    base_cols = [c for c in wl.columns if c in merged.columns]
    extra_cols = [c for c in merged.columns if c not in base_cols]
    ordered = merged[base_cols + extra_cols]

    out_path = out_dir / f"worklogs_with_rates_{ts}.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        ordered.to_excel(writer, index=False, sheet_name=SHEET_NAME_OUT)

    # строго копируем исходный столбец "Ссылка на задачу"
    copy_original_task_links(wl_path, out_path, dst_sheet=SHEET_NAME_OUT)

    print("[✓] Готово:", out_path)
    return out_path


if __name__ == "__main__":
    main()
